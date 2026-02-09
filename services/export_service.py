"""
Export Service
Excel va PDF export funksiyalari
"""
import os
from datetime import date, datetime
from typing import List, Dict
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.chart import BarChart, Reference

class ExportService:
    """Export xizmatlari"""
    
    def __init__(self):
        self.export_dir = "exports"
        os.makedirs(self.export_dir, exist_ok=True)
    
    async def export_to_excel(
        self, 
        user_id: int, 
        transactions: List[Dict],
        stats: Dict,
        month: str = None
    ) -> str:
        """
        Excel fayl yaratish
        
        Returns:
            Fayl yo'li
        """
        if month is None:
            month = date.today().strftime("%Y-%m")
        
        filename = f"aris_hisobot_{user_id}_{month}.xlsx"
        filepath = os.path.join(self.export_dir, filename)
        
        # Workbook yaratish
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Hisobot"
        
        # Sarlavha
        ws['A1'] = "ARIS - Moliyaviy Hisobot"
        ws['A1'].font = Font(size=16, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center')
        ws.merge_cells('A1:E1')
        
        # Umumiy statistika
        ws['A3'] = "Umumiy Statistika"
        ws['A3'].font = Font(size=14, bold=True)
        
        ws['A4'] = "Jami Kirim:"
        ws['B4'] = stats.get('total_income', 0)
        ws['B4'].number_format = '#,##0'
        
        ws['A5'] = "Jami Chiqim:"
        ws['B5'] = stats.get('total_expense', 0)
        ws['B5'].number_format = '#,##0'
        
        ws['A6'] = "Balans:"
        ws['B6'] = stats.get('balance', 0)
        ws['B6'].number_format = '#,##0'
        ws['B6'].font = Font(bold=True)
        
        # Tranzaksiyalar jadvali
        ws['A8'] = "Tranzaksiyalar"
        ws['A8'].font = Font(size=14, bold=True)
        
        # Jadval sarlavhalari
        headers = ['Sana', 'Turi', 'Summa', 'Kategoriya', 'Tavsif']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=9, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Tranzaksiyalar
        for row, trans in enumerate(transactions, start=10):
            ws.cell(row=row, column=1, value=trans.get('date', ''))
            ws.cell(row=row, column=2, value=trans.get('type', ''))
            ws.cell(row=row, column=3, value=trans.get('amount', 0))
            ws.cell(row=row, column=3).number_format = '#,##0'
            ws.cell(row=row, column=4, value=trans.get('category', ''))
            ws.cell(row=row, column=5, value=trans.get('description', ''))
        
        # Ustun kengligini sozlash
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 10
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 30
        
        # Saqlash
        wb.save(filepath)
        
        return filepath
    
    async def export_to_pdf(
        self,
        user_id: int,
        transactions: List[Dict],
        stats: Dict,
        month: str = None
    ) -> str:
        """
        PDF fayl yaratish (oddiy versiya)
        
        Returns:
            Fayl yo'li
        """
        if month is None:
            month = date.today().strftime("%Y-%m")
        
        filename = f"aris_hisobot_{user_id}_{month}.txt"
        filepath = os.path.join(self.export_dir, filename)
        
        # Hozircha oddiy text fayl (keyinroq PDF qilamiz)
        with open(filepath, 'w', encoding='utf-8') as f:
            f.write("=" * 50 + "\n")
            f.write("ARIS - Moliyaviy Hisobot\n")
            f.write(f"Sana: {datetime.now().strftime('%Y-%m-%d %H:%M')}\n")
            f.write("=" * 50 + "\n\n")
            
            f.write("UMUMIY STATISTIKA\n")
            f.write("-" * 50 + "\n")
            f.write(f"Jami Kirim:  {stats.get('total_income', 0):,} so'm\n")
            f.write(f"Jami Chiqim: {stats.get('total_expense', 0):,} so'm\n")
            f.write(f"Balans:      {stats.get('balance', 0):,} so'm\n")
            f.write("\n")
            
            f.write("TRANZAKSIYALAR\n")
            f.write("-" * 50 + "\n")
            for trans in transactions:
                f.write(f"{trans.get('date')} | {trans.get('type'):7} | "
                       f"{trans.get('amount'):>10,} | {trans.get('category')}\n")
                if trans.get('description'):
                    f.write(f"  {trans.get('description')}\n")
                f.write("\n")
        
        return filepath
